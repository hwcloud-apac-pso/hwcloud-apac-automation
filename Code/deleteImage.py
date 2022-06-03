# coding: utf-8

from huaweicloudsdkcore.auth.credentials import BasicCredentials
from huaweicloudsdkims.v2.region.ims_region import ImsRegion
from huaweicloudsdkcore.exceptions import exceptions
from huaweicloudsdkims.v2 import *
from dotenv import load_dotenv
import os
import pandas as pd
from time import sleep, perf_counter
from openpyxl import load_workbook
from utils import declare_client

pd.options.mode.chained_assignment = None


if __name__ == "__main__":

    load_dotenv("config.env")
    ak = os.getenv("ACCESSKEY")
    sk = os.getenv("SECRETKEY")
    resource_region = "ap-southeast-3"

    #declare clients
    clients = declare_client(resource_region, ak, sk)
    
    
    #resource list path
    path = "resource_list.xlsx"
    
    #load workboo
    wb = load_workbook(filename = path)  
    source = wb['Image']

    df = pd.read_excel(path, 'Image', header=None)
    df = df[0].values.tolist()

    for count,img_id in enumerate(df):
        try:
            print("Start to delete image ......")
            request = GlanceDeleteImageRequest(image_id = img_id)
            request.body = GlanceDeleteImageRequestBody(delete_backup = True)
            response = clients[0].glance_delete_image(request).to_dict()
            print(f"Image {img_id} deleted")
            source.cell(row=count+1, column=1).value = None
        except exceptions.ClientRequestException as e:
            print(e.status_code)
            print(e.request_id)
            print(e.error_code)
            print(e.error_msg)
    
    wb.save(path)
    
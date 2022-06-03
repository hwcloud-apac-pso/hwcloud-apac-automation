# coding: utf-8
from huaweicloudsdkcore.auth.credentials import BasicCredentials
from huaweicloudsdkims.v2.region.ims_region import ImsRegion
from huaweicloudsdkecs.v2.region.ecs_region import EcsRegion
from huaweicloudsdkcore.exceptions import exceptions
from huaweicloudsdkims.v2 import *
from huaweicloudsdkecs.v2 import *
from dotenv import load_dotenv
import os
import pandas as pd
import sys
from time import sleep, perf_counter
pd.options.mode.chained_assignment = None
from openpyxl import load_workbook
from utils import create_image, create_ecs, declare_client

if __name__ == "__main__":
    load_dotenv("config.env")
    ak = os.getenv("ACCESSKEY")
    sk = os.getenv("SECRETKEY")
    adminPW = os.getenv("ADMINPW")
    resource_region = "ap-southeast-3"
    
    # specify the network id
    vpcID = os.getenv("vpcID")
    subnetID = os.getenv("subnetID")
    
    
    #declare clients
    clients = declare_client(resource_region, ak, sk)
    
    # load excel workbook with specific worksheet number
    f = open('num.txt','r')
    ws_num = int(f.read())
    df = pd.read_excel('ids_list.xlsx', sheet_name = f"Sheet{ws_num}")
    # total rows in worksheet
    row = len(df.index)
    
    #resource list path
    path = "resource_list.xlsx"
    
    #get start counter for success_list
    success_list = pd.read_excel(path, 'Image', header=None)
    success_start = len(success_list)
    
    #get start counter for fail_list
    fail_list = pd.read_excel(path, 'Failed')
    #length +1 to start from row two, row one is header 
    fail_start = len(fail_list) + 1
    
    #load workbook
    wb = load_workbook(filename = path)
    
    #init counter, 
    count = 0
    success_counter = 1
    fail_counter = 1
    
    #loop
    while count < row:
        # obtain info from excel file
        vaultID = df['vaultId'].iloc[count]
        backupID = df['backupId'].iloc[count]
        flavorREF = df['flavorRef'].iloc[count]
        eipID = df['eipId'].iloc[count]
        smsVERSION = df['sms_version'].iloc[count]
        
        if success_start < 10:
            ecs_name = f"ecs-00{success_start}"
            img_name = f"img-00{success_start}"
        else:
            ecs_name = f"ecs-0{success_start}"
            img_name = f"img-0{success_start}"
 
        img_id = None
        count += 1
        run = True
        while run:
            try:
                ##### CREATE FULL ECS IMAGE FROM CBR BACKUPS #####
                # create image from CBR backups
                img_id = create_image(clients[0], img_name, backupID, vaultID)

                if img_id is not None:
                    ##### PROVISION ECS USING CREATED IMAGE #####
                    # create ecs using created image
                    ecs_id = create_ecs(clients[1], smsVERSION, eipID, subnetID, img_id, flavorREF, ecs_name, vpcID, adminPW)
                else:
                    raise ValueError('throw error')
                if img_id is not None and ecs_id is not None:
                    # add server id to resource_list.xlsx
                    source = wb['Server']
                    source.cell(row=success_start+success_counter, column=1).value = ecs_id

                    # add image id to resource_list.xlsx
                    source = wb['Image']
                    source.cell(row=success_start+success_counter, column=1).value = img_id
                    success_counter += 1
                    run = False  
                else:
                    raise ValueError('throw error')
            except exceptions.ClientRequestException as e:
                if img_id is not None:
                    source = wb["Failed"]
                    source.cell(row=fail_start+fail_counter, column=1).value = backupID
                    source.cell(row=fail_start+fail_counter, column=2).value = vaultID
                    source.cell(row=fail_start+fail_counter, column=3).value = img_id
                    fail_counter += 1
                else:
                    source = wb["Failed"]
                    source.cell(row=fail_start+fail_counter, column=1).value = backupID
                    source.cell(row=fail_start+fail_counter, column=2).value = vaultID
                    source.cell(row=fail_start+fail_counter, column=3).value = None
                    fail_counter += 1
                print(e.status_code)
                print(e.request_id)
                print(e.error_code)
                print(e.error_msg)
                run = False

    wb.save(path)
# coding: utf-8

from huaweicloudsdkcore.auth.credentials import BasicCredentials
from huaweicloudsdkecs.v2.region.ecs_region import EcsRegion
from huaweicloudsdkcore.exceptions import exceptions
from huaweicloudsdkecs.v2 import *
from dotenv import load_dotenv
import os
import pandas as pd
from time import sleep, perf_counter
from openpyxl import load_workbook
from utils import declare_client

pd.options.mode.chained_assignment = None

if __name__ == "__main__":

    load_dotenv("config.env")
    ak = os.getenv("ACCESSKEY")
    sk = os.getenv("SECRETKEY")
    resource_region = "ap-southeast-3"
    
    #declare clients
    clients = declare_client(resource_region, ak, sk)
    
    #resource list path
    path = "resource_list.xlsx"
    
    #load workbook
    wb = load_workbook(filename = path)  
    source = wb['Server']

    df = pd.read_excel(path, 'Server', header=None)
    df = df[0].values.tolist()

    for count,server_id in enumerate(df):
        try:
            print("Start to delete server ......")
            request = DeleteServersRequest()
            listServerIdServersbody = [
                ServerId(
                    id = server_id
                )
            ]
            request.body = DeleteServersRequestBody(
                servers=listServerIdServersbody,
                delete_volume = True,
                delete_publicip = False
            )
            response = clients[1].delete_servers(request).to_dict()
            deleteecs_job_id = response['job_id']

            # check the status of ecs deletion
            while True:
                showecsjob_req = ShowJobRequest(job_id = deleteecs_job_id)
                showecsjob_res = clients[1].show_job(showecsjob_req).to_dict()
                showecsjob_status = showecsjob_res['status']
                print("ECS Status -----> " + showecsjob_status)
                sleep(10)
                if showecsjob_status == "SUCCESS":
                    print("ECS with ID " + server_id + " deleted successfully!\n")
                    source.cell(row=count+1, column=1).value = None
                    break
            
        except exceptions.ClientRequestException as e:
            print(e.status_code)
            print(e.request_id)
            print(e.error_code)
            print(e.error_msg)

    wb.save(path)